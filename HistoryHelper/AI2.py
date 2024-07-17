import re
from openai import OpenAI
import openpyxl
from datetime import datetime
import os 
import tkinter as tk
import threading
import json  # Import the json module
import tkinter.filedialog as filedialog
import webbrowser
from tkinter import ttk
import time

file_path=""
api_key=""

stop_processing = False
waiting_for_user = False
wait=False

# Function to load configuration from a JSON file
def toggle_configuration():
    if file_path_label.winfo_ismapped():
        # Hide configuration widgets
        file_path_entry.delete(0, tk.END)
        file_path_label.grid_remove()
        file_path_entry.grid_remove()

        api_key_entry.delete(0, tk.END)
        api_key_label.grid_remove()
        api_key_entry.grid_remove()
        update_button.grid_remove()
        select_file_button.grid_remove()
        open_openai_button.grid_remove()
    else:
        # Show configuration widgets
        file_path_label.grid()
        file_path_entry.grid()
        file_path_entry.insert(tk.END, file_path)
        select_file_button.grid()
        api_key_label.grid()
        api_key_entry.grid()
        api_key_entry.insert(tk.END, api_key)
        update_button.grid()
        open_openai_button.grid()
        
        
def open_openai_website():
    webbrowser.open("https://platform.openai.com/api-keys")
    
def select_file_path():
    global file_path
    file_path = filedialog.askopenfilename()
    file_path_entry.delete(0, tk.END)
    file_path_entry.insert(tk.END, file_path)
    
def create_prompt(event_name):
    known_events = load_historical_events()
    known_events_str = ", ".join(known_events)
    prompt = f"I already know about these events: {known_events_str}. Write me about event {event_name}, assuming I know about the events listed above."
    return prompt

    
def load_historical_events():
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    events = []
    for row in sheet.iter_cols(min_row=1, max_row=1, min_col=2, values_only=True):
        events.extend(row)
    workbook.close()
    return events

def load_configuration():
    global file_path, api_key
    # Get the directory of the current script
    script_dir = os.path.dirname(os.path.realpath(__file__))
    config_file_path = os.path.join(script_dir, "config.json")
    try:
        with open(config_file_path, "r") as f:
            config = json.load(f)
            file_path = config.get("file_path", "")
            api_key = config.get("api_key", "")
    except FileNotFoundError:
        print("Configuration file not found.")
        # You can set default values here if needed

# Function to save configuration to a JSON file
def save_configuration():
    global file_path, api_key
    # Get the directory of the current script
    script_dir = os.path.dirname(os.path.realpath(__file__))
    config_file_path = os.path.join(script_dir, "config.json")
    config = {"file_path": file_path, "api_key": api_key}
    with open(config_file_path, "w") as f:
        json.dump(config, f)

# Load configuration when the script starts
load_configuration()

# Function to update file path and API key
def update_configuration():
    global file_path, api_key
    file_path = file_path_entry.get()
    api_key = api_key_entry.get()
    save_configuration()

stop_processing = False

def quit_processing():
    global stop_processing
    if stop_processing==False:
        stop_processing = True
        output_text.insert(tk.END, "Quit processing successful\n")
        output_text.update_idletasks()
    
def start_processing_thread():
    processing_thread = threading.Thread(target=process_events)
    processing_thread.start()

def get_response(prompt):
    client = OpenAI(api_key=api_key)

    completion = client.chat.completions.create(
    model="gpt-3.5-turbo",
    messages=[
        {"role": "system", "content": "In this chat if I input the Historical event, I want you to solely output the date of the event strictly in the format of xx.yy.zzzz or xx.yy.zzzz - qq.ww.eeee, what people were involved in the event, which countries were involved or where it happened, what is the event, and why it happened. For what and why questions you must write one paragraph of a text for each. For what section only write everything about what the event is and whhat were the consequences of it. For why section only write the causes of the event, and why it happened. Do not write consequences in the why section. Write the text with as simple words as possible, while still being specific. When there are multiple dates, such as for the Hiroshima and Nagasaki bombings, please write them in the following format: xx.yy.zzzz - qq.ww.eeee, where xx.yy.zzzz is the first date and qq.ww.eeee is the second date. If there are more than two dates, use the same format, but write the first date - last date. If you do not know the specific date and only know the year of the event, write it in the format 01.01.xxxx. The date should always be in one of two formats: xx.yy.zzzz or xx.yy.zzzz - qq.ww.eeee. If you are unsure of the specific year, write it as a range of possible dates, like xx.yy.zzzz - qq.ww.eeee. Always remember: the date should always be in the format of xx.yy.zzzz or xx.yy.zzzz - qq.ww.eeee. I will also provide you with historical events I already know, so you shouldn't explain again what are the events or what do terms mean. Perfect examples are: Anschluss\nYour Output:\nDate: 12.03.1938\nPeople Involved: Adolf Hitler, Kurt Schuschnigg\nCountries Involved: Germany, Austria\nEvent: The annexation of Austria by Nazi Germany. It was a major step in Hitler's plan to create a Greater German Reich. He pressured the Austrian government to appoint a pro-German chancellor, and he threatened military action if Austria resisted\nWhy it happened: The Anschluss had deep historical and political roots. Following World War I, the Treaty of Versailles prohibited the union of Germany and Austria. However, the idea of a Greater Germany, incorporating all German-speaking people, gained popularity in Nazi ideology. Hitler, with expansionist goals, aimed to unite all German-speaking regions. The Anschluss also capitalized on the internal political struggles within Austria, allowing Hitler to orchestrate the annexation through both diplomatic pressure and military intervention, solidifying his influence in the region.\n User:\nRhineland remilitarisation\nYour Output:\nDate: 7.03.1936\nPeople Involved: Adolf Hitler\nCountries Involved: Germany, France\nEvent: The remilitarization of the Rhineland, a demilitarized zone in western Germany, was a bold move by Adolf Hitler that violated the Treaty of Versailles, which had ended World War I.\nWhy it happened: Hitler was determined to restore Germany to its former glory and believed that rearming the country was essential to achieving this goal. He also saw the Rhineland as a buffer zone between Germany and France, and he wanted to remove any obstacles to German expansion."},
        {"role": "user", "content": prompt}
    ]
    )
    text=completion.choices[0].message.content
    return text

def find_date(text):
    pattern = r'Date:\s*(.*?)\n'
    date_format = "%d.%m.%Y"

    date_match = re.search(pattern, text)
    if date_match:
        date = date_match.group(1)
        try:
            # Check if the date matches the expected format
            if "-" in date:
                init_date, end_date = date.split(" - ")
                datetime.strptime(init_date.strip(), date_format)
                datetime.strptime(end_date.strip(), date_format)
            else:
                datetime.strptime(date.strip(), date_format)
            return date
        except:
            return f"01.01.{date}"
    else:
        print("Date not found")
        return None

        
def find_people(text):
    pattern = r'People Involved:\s*(.*?)\n'
    people_match = re.search(pattern, text)
    if people_match:
        people_value = people_match.group(1)
        names = people_value.split(', ')
        return ', '.join(names)
    else:
        print("People Involved not found")
        
def find_countries(text):
    pattern = r'Countries Involved:\s*(.*?)\n'

    countries_match = re.search(pattern, text)
    if countries_match:
        countries_value = countries_match.group(1)
        country_names = countries_value.split(', ')
        return ', '.join(country_names)
    else:
        print("Countries Involved not found")
        
def find_what(text):
    pattern = r'Event:\s*(.*?)\n'

    event_match = re.search(pattern, text)
    if event_match:
        event_value = event_match.group(1)
        return str(event_value)
    else:
        print("Event not found")
        
def find_why(text):
    pattern = r'Why it happened:\s*(.*?)(?:\n|$)'
    why_match = re.search(pattern, text)

    if why_match:
        reason = why_match.group(1)
        return str(reason)
    else:
        print("Why not found")
        
def insert_event(sheet, col, event, when, who, where, what, why):
    sheet.cell(row=1, column=col+1).value = event
    sheet.cell(row=2, column=col+1).value = when
    sheet.cell(row=3, column=col+1).value = who
    sheet.cell(row=4, column=col+1).value = where
    sheet.cell(row=5, column=col+1).value = what
    sheet.cell(row=6, column=col+1).value = why

def wrap_text(sheet):
        for row in sheet.iter_rows():
            for cell in row:
                # Wrap text in the cell
                cell.alignment = openpyxl.styles.Alignment(wrapText=True)

def find_init_date(full_date):
    date_format = "%d.%m.%Y"
    if "-" in full_date:
        init_date, _ = full_date.split(" - ")
        init_date = datetime.strptime(init_date, date_format)
    else:
        init_date = datetime.strptime(full_date, date_format)
    return init_date

current_event_details = {}

def apply_changes():
    global current_event_details, wait
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    event_name = event_name_text.get("1.0", tk.END)
    event_date = date_text.get("1.0", tk.END).strip()
    event_people = people_text.get("1.0", tk.END)
    event_countries = countries_text.get("1.0", tk.END)
    event_what = what_text.get("1.0", tk.END)
    event_why = why_text.get("1.0", tk.END)

    date_to_fit = find_init_date(event_date)
    col = 2

    while True:
        start_year_col = str(sheet.cell(row=2, column=col).value)
        date1 = find_init_date(start_year_col)
        end_year_col = str(sheet.cell(row=2, column=col + 1).value)
        date2 = find_init_date(end_year_col)

        if date1 < date_to_fit < date2 or date1 == date_to_fit:
            sheet.insert_cols(col + 1, amount=1)
            insert_event(sheet, col, event_name, event_date, event_people, event_countries, event_what, event_why)
            break
        elif date_to_fit < date1:
            sheet.insert_cols(col, amount=1)
            insert_event(sheet, col - 1, event_name, event_date, event_people, event_countries, event_what, event_why)
            break
        elif date_to_fit > date2:
            empty_cells = True
            for column in sheet.iter_cols(min_row=2, max_row=2, min_col=col + 2):
                for cell in column:
                    if cell.value is not None:
                        empty_cells = False
                        break
                if not empty_cells:
                    break
            if empty_cells:
                sheet.insert_cols(col + 2, amount=1)
                insert_event(sheet, col+1, event_name, event_date, event_people, event_countries, event_what, event_why)
                break
        col += 1
    wrap_text(sheet)
    workbook.save(file_path)
    workbook.close()
    output_text.insert(tk.END, f"Changes applied to {event_name} 🆗\n")

    # Hide the "Apply Changes" button and clear event information
    apply_changes_button.grid_remove()
    event_name_text.delete("1.0", tk.END)
    date_text.delete("1.0", tk.END)
    people_text.delete("1.0", tk.END)
    countries_text.delete("1.0", tk.END)
    what_text.delete("1.0", tk.END)
    why_text.delete("1.0", tk.END)
    
    wait=False


def process_events():
    global stop_processing, waiting_for_user, current_event_details, wait
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    event_names = event_names_entry.get().split("; ")
    event_names_entry.delete(0, tk.END)
    output_text.delete("1.0", tk.END)
    for event_name in event_names:
        event_done = False
        unsuccessful_trials = 0
        while (not event_done or unsuccessful_trials >= 20) and not stop_processing:
            try:
                text = get_response(create_prompt(event_name))
                
                # Clear previous results
                event_name_text.delete("1.0", tk.END)
                date_text.delete("1.0", tk.END)
                people_text.delete("1.0", tk.END)
                countries_text.delete("1.0", tk.END)
                what_text.delete("1.0", tk.END)
                why_text.delete("1.0", tk.END)

                event_name_text.insert(tk.END, event_name)

                event_date = find_date(text).strip()
                event_people = find_people(text)
                event_countries = find_countries(text)
                event_what = find_what(text)
                event_why = find_why(text)

                # Insert new results
                date_text.insert(tk.END, event_date)
                people_text.insert(tk.END, event_people)
                countries_text.insert(tk.END, event_countries)
                what_text.insert(tk.END, event_what)
                why_text.insert(tk.END, event_why)

                event_done = True
                output_text.insert(tk.END, f"{event_name} ✅\n")
                output_text.update_idletasks()
            except Exception as e:
                unsuccessful_trials += 1
                output_text.insert(tk.END, f"{unsuccessful_trials}th unsuccessful trial for event {event_name}. ❌ Wait for more attempts or try a different event.\n")
                output_text.update_idletasks()
                if unsuccessful_trials >= 20:
                    output_text.insert(tk.END, f"Too many unsuccessful trials for event {event_name}. 🛑\n")
                    output_text.update_idletasks()
                    break
        apply_changes_button.grid()
        wait=True
        while wait:
            time.sleep(0)
    output_text.insert(tk.END, f"Finished generating 🆗\n")
    stop_processing = False
    workbook.close()



window = tk.Tk()
window.title("History Helper")

# Create input fields for file path and API key
file_path_label = ttk.Label(window, text="File Path:")
file_path_entry = ttk.Entry(window, width=50)
api_key_label = ttk.Label(window, text="API Key:")
api_key_entry = ttk.Entry(window, width=50)
update_button = ttk.Button(window, text="Update Configuration", command=update_configuration)
select_file_button = ttk.Button(window, text="Select Excel File", command=select_file_path)
open_openai_button = ttk.Button(window, text="Get OpenAI API Key", command=open_openai_website)
event_names_entry = ttk.Entry(window)
process_button = ttk.Button(window, text="Process Events", command=start_processing_thread)
output_text = tk.Text(window, height=5, width=50)
quit_button = ttk.Button(window, text="Quit Processing", command=quit_processing)
toggle_button = ttk.Button(window, text="Toggle Configuration", command=toggle_configuration)

event_name_label = ttk.Label(window, text="Event Name:")
event_name_text = tk.Text(window, height=1, width=50)
date_label = ttk.Label(window, text="Date:")
date_text = tk.Text(window, height=1, width=50)
people_label = ttk.Label(window, text="People Involved:")
people_text = tk.Text(window, height=1, width=50)
countries_label = ttk.Label(window, text="Countries Involved:")
countries_text = tk.Text(window, height=1, width=50)
what_label = ttk.Label(window, text="What:")
what_text = tk.Text(window, height=3, width=50)
why_label = ttk.Label(window, text="Why:")
why_text = tk.Text(window, height=3, width=50)

event_name_label.grid(row=5, column=0, padx=5, pady=5, sticky="W")
event_name_text.grid(row=5, column=1, padx=5, pady=5, sticky="EW", columnspan=2)
date_label.grid(row=6, column=0, padx=5, pady=5, sticky="W")
date_text.grid(row=6, column=1, padx=5, pady=5, sticky="EW", columnspan=2)
people_label.grid(row=7, column=0, padx=5, pady=5, sticky="W")
people_text.grid(row=7, column=1, padx=5, pady=5, sticky="EW", columnspan=2)
countries_label.grid(row=8, column=0, padx=5, pady=5, sticky="W")
countries_text.grid(row=8, column=1, padx=5, pady=5, sticky="EW", columnspan=2)
what_label.grid(row=9, column=0, padx=5, pady=5, sticky="W")
what_text.grid(row=9, column=1, padx=5, pady=5, sticky="EW", columnspan=2)
why_label.grid(row=10, column=0, padx=5, pady=5, sticky="W")
why_text.grid(row=10, column=1, padx=5, pady=5, sticky="EW", columnspan=2)

apply_changes_button = ttk.Button(window, text="Apply Changes", command=apply_changes)
apply_changes_button.grid(row=11, column=0, padx=5, pady=5, sticky="EW")
apply_changes_button.grid_remove()


# Adjust grid layout for better spacing and resizing
file_path_label.grid(row=0, column=0, padx=5, pady=5, sticky="W")
file_path_entry.grid(row=0, column=1, padx=5, pady=5, sticky="EW")
select_file_button.grid(row=0, column=2, padx=5, pady=5, sticky="EW")
api_key_label.grid(row=1, column=0, padx=5, pady=5, sticky="W")
api_key_entry.grid(row=1, column=1, padx=5, pady=5, sticky="EW")
update_button.grid(row=2, column=2, padx=5, pady=5, sticky="EW")
open_openai_button.grid(row=1, column=2, padx=5, pady=5, sticky="EW")
event_names_entry.grid(row=2, column=0, columnspan=2, padx=5, pady=5, sticky="EW")
process_button.grid(row=3, column=0, padx=5, pady=5, sticky="EW")
output_text.grid(row=4, column=1, padx=5, pady=5, sticky="EW")
quit_button.grid(row=4, column=0, padx=5, pady=5, sticky="EW")
toggle_button.grid(row=4, column=0, padx=5, pady=5, sticky="W")

# Make columns and rows resize correctly
window.grid_columnconfigure(1, weight=1)
window.grid_columnconfigure(3, weight=1)
window.grid_rowconfigure(4, weight=1)

window.after(100, toggle_button.invoke)
window.mainloop()
